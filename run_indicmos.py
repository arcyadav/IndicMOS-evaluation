import argparse
import csv
import statistics
from collections import defaultdict
from pathlib import Path

import torch
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from run_single_indicmos import HF_PATH, LANG_ID_MAPPING, load_audio_16k_mono, load_model


AUDIO_EXTENSIONS = {".wav", ".mp3", ".flac", ".ogg", ".m4a", ".aac"}

FOLDER_LANG_HINTS = {
    "hindi": "hi",
    "telugu": "te",
    "kannada": "kn",
    "marathi": "mr",
    "bengali": "bn",
    "english": "en",
    "chhattisgarhi": "ch",
}


def infer_langid(path: Path) -> str | None:
    folder_name = path.parent.name.lower()
    for hint, langid in FOLDER_LANG_HINTS.items():
        if hint in folder_name:
            return langid
    return None


def iter_audio_files(input_dir: Path) -> list[Path]:
    audio_files = [
        path
        for path in input_dir.rglob("*")
        if path.is_file() and path.suffix.lower() in AUDIO_EXTENSIONS
    ]
    return sorted(audio_files, key=lambda path: path.as_posix().lower())


def score_audio(model, audio_path: Path, device: str, langid: str | None, use_langid: bool) -> float:
    wav = load_audio_16k_mono(audio_path).to(device)
    lengths = torch.tensor([wav.shape[-1]], dtype=torch.long, device=device)
    lang_tensor = None

    if use_langid:
        if langid is None:
            raise ValueError("Could not infer language id")
        lang_tensor = torch.tensor([LANG_ID_MAPPING[langid]], dtype=torch.long, device=device)

    with torch.no_grad():
        score = model(
            wav,
            lang_data=lang_tensor,
            lengths=lengths,
            batch_mode=True,
        ).squeeze().cpu().item()

    return max(1.0, min(5.0, float(score)))


def write_file_scores(rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["folder", "file", "relative_path", "langid", "mos", "status", "error"]
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_folder_summary(rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = ["folder", "num_files", "num_scored", "avg_mos", "min_mos", "max_mos", "failed"]
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_excel_report(file_rows: list[dict], summary_rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4", end_color="4472C4")
    body_font = Font(name="Arial")
    center = Alignment(horizontal="center")

    # --- Folder summary sheet ---
    summary_sheet = wb.active
    summary_sheet.title = "Folder Summary"
    summary_fields = ["folder", "num_files", "num_scored", "avg_mos", "min_mos", "max_mos", "failed"]
    summary_headers = ["Folder", "Num Files", "Num Scored", "Avg MOS", "Min MOS", "Max MOS", "Failed"]

    summary_sheet.append(summary_headers)
    for cell in summary_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row in summary_rows:
        values = []
        for field in summary_fields:
            value = row[field]
            if field in {"avg_mos", "min_mos", "max_mos"} and value != "":
                value = float(value)
            values.append(value)
        summary_sheet.append(values)

    for row_idx in range(2, summary_sheet.max_row + 1):
        for col_idx in range(1, len(summary_headers) + 1):
            cell = summary_sheet.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.alignment = center
            if summary_headers[col_idx - 1] in {"Avg MOS", "Min MOS", "Max MOS"}:
                cell.number_format = "0.0000"

    for col_idx, header in enumerate(summary_headers, start=1):
        summary_sheet.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) + 4)
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions

    # --- File scores sheet ---
    file_sheet = wb.create_sheet("File Scores")
    file_fields = ["folder", "file", "relative_path", "langid", "mos", "status", "error"]
    file_headers = ["Folder", "File", "Relative Path", "Lang ID", "MOS", "Status", "Error"]

    file_sheet.append(file_headers)
    for cell in file_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    error_fill = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

    for row in file_rows:
        values = []
        for field in file_fields:
            value = row[field]
            if field == "mos" and value != "":
                value = float(value)
            values.append(value)
        file_sheet.append(values)

    for row_idx in range(2, file_sheet.max_row + 1):
        status_value = file_sheet.cell(row=row_idx, column=file_headers.index("Status") + 1).value
        for col_idx in range(1, len(file_headers) + 1):
            cell = file_sheet.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            if file_headers[col_idx - 1] == "MOS":
                cell.number_format = "0.0000"
                cell.alignment = center
            if status_value == "error":
                cell.fill = error_fill

    for col_idx, header in enumerate(file_headers, start=1):
        width = 18 if header != "Relative Path" else 40
        file_sheet.column_dimensions[get_column_letter(col_idx)].width = width
    file_sheet.freeze_panes = "A2"
    file_sheet.auto_filter.ref = file_sheet.dimensions

    wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Score a folder tree with IndicMOS.")
    parser.add_argument("--input-dir", type=Path, default=Path("MedhaApp_TTS_Audio"))
    parser.add_argument("--output-dir", type=Path, default=Path("results") / "indicmos")
    parser.add_argument("--device", default="cpu", help="cpu or cuda")
    parser.add_argument("--model-cache", default=HF_PATH, help="Hugging Face model cache folder")
    parser.add_argument("--use-langid", action="store_true", help="Use language-aware IndicMOS model")
    parser.add_argument("--limit", type=int, default=None, help="Optional quick-test limit")
    args = parser.parse_args()

    if not args.input_dir.exists():
        raise FileNotFoundError(args.input_dir)

    audio_files = iter_audio_files(args.input_dir)
    if args.limit is not None:
        audio_files = audio_files[: args.limit]

    if not audio_files:
        raise ValueError(f"No audio files found under {args.input_dir}")

    model = load_model(
        use_cer=False,
        use_langid=args.use_langid,
        download_path=args.model_cache,
        device=args.device,
    )

    file_rows = []
    by_folder = defaultdict(list)
    folder_totals = defaultdict(int)

    for idx, audio_path in enumerate(audio_files, start=1):
        rel_path = audio_path.relative_to(args.input_dir)
        folder = rel_path.parts[0] if len(rel_path.parts) > 1 else "."
        folder_totals[folder] += 1
        langid = infer_langid(audio_path)

        print(f"[{idx}/{len(audio_files)}] {rel_path}")

        try:
            mos = score_audio(model, audio_path, args.device, langid, args.use_langid)
            mos_text = f"{mos:.4f}"
            status = "ok"
            error = ""
            by_folder[folder].append(mos)
        except Exception as exc:
            mos_text = ""
            status = "error"
            error = str(exc)

        file_rows.append(
            {
                "folder": folder,
                "file": audio_path.name,
                "relative_path": rel_path.as_posix(),
                "langid": langid or "",
                "mos": mos_text,
                "status": status,
                "error": error,
            }
        )

    summary_rows = []
    for folder in sorted(folder_totals):
        scores = by_folder[folder]
        failed = folder_totals[folder] - len(scores)
        summary_rows.append(
            {
                "folder": folder,
                "num_files": folder_totals[folder],
                "num_scored": len(scores),
                "avg_mos": f"{statistics.mean(scores):.4f}" if scores else "",
                "min_mos": f"{min(scores):.4f}" if scores else "",
                "max_mos": f"{max(scores):.4f}" if scores else "",
                "failed": failed,
            }
        )

    file_scores_path = args.output_dir / "indicmos_file_scores.csv"
    summary_path = args.output_dir / "indicmos_folder_summary.csv"
    excel_path = args.output_dir / "indicmos_report.xlsx"
    write_file_scores(file_rows, file_scores_path)
    write_folder_summary(summary_rows, summary_path)
    write_excel_report(file_rows, summary_rows, excel_path)

    print(f"Saved file scores: {file_scores_path}")
    print(f"Saved folder summary: {summary_path}")
    print(f"Saved Excel report: {excel_path}")


if __name__ == "__main__":
    main()
