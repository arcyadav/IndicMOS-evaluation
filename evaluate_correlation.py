"""
Extract human MOS and Predicted MOS rows from the Google Form responses
Excel sheet, and compute correlation / error metrics comparing them.

Usage:
    python evaluate_correlation.py path/to/Responses.xlsx

By default, the output CSV is written to:
    results/mos_correlation_results.csv

"""

import argparse
import csv
from pathlib import Path

import numpy as np
import openpyxl
from scipy import stats

# EDIT THIS to match the left-to-right order of audio blocks in your sheet.
LANGUAGES = ["English", "Hindi", "Kannada", "Telugu"]
PROJECT_DIR = Path(__file__).resolve().parent


def find_summary_row(sheet, label: str) -> list:
    """Return the full row of values whose first cell equals `label`."""
    for row in sheet.iter_rows(values_only=True):
        first_cell = str(row[0]).strip() if row[0] is not None else ""
        if first_cell == label:
            return list(row)
    raise ValueError(f"Could not find a row starting with '{label}'")


def find_audio_blocks(header_row: list) -> list[list[int]]:
    """Group consecutive column indices whose header mentions 'Audio'."""
    blocks = []
    current = []
    for idx, header in enumerate(header_row):
        is_audio_col = header is not None and "Audio" in str(header)
        if is_audio_col:
            current.append(idx)
        elif current:
            blocks.append(current)
            current = []
    if current:
        blocks.append(current)
    return blocks


def metrics(human: np.ndarray, pred: np.ndarray) -> dict:
    pearson_r, pearson_p = stats.pearsonr(human, pred)
    spearman_r, spearman_p = stats.spearmanr(human, pred)
    kendall_r, kendall_p = stats.kendalltau(human, pred)
    return {
        "n": len(human),
        "pearson_r": pearson_r, "pearson_p": pearson_p,
        "spearman_rho": spearman_r, "spearman_p": spearman_p,
        "kendall_tau": kendall_r, "kendall_p": kendall_p,
        "mae": np.mean(np.abs(human - pred)),
        "rmse": np.sqrt(np.mean((human - pred) ** 2)),
    }


def main():
    parser = argparse.ArgumentParser(description="Compare human vs predicted MOS.")
    parser.add_argument("excel_path", type=Path)
    parser.add_argument("--output-csv", type=Path, default=None,
                         help="Defaults to results/mos_correlation_results.csv")
    args = parser.parse_args()

    args.excel_path = args.excel_path.resolve()
    if args.output_csv is None:
        args.output_csv = PROJECT_DIR / "results" / "mos_correlation_results.csv"
    args.output_csv = args.output_csv.resolve()

    wb = openpyxl.load_workbook(args.excel_path, data_only=True)
    sheet = wb.worksheets[0]

    header_row = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    human_row = find_summary_row(sheet, "MOS")
    pred_row = find_summary_row(sheet, "Predicted MOS")

    blocks = find_audio_blocks(header_row)
    if len(blocks) != len(LANGUAGES):
        raise ValueError(
            f"Found {len(blocks)} audio blocks in the sheet but LANGUAGES has "
            f"{len(LANGUAGES)} entries. Update the LANGUAGES list to match."
        )

    all_human, all_pred, results = [], [], []
    for lang, cols in zip(LANGUAGES, blocks):
        human = np.array([human_row[c] for c in cols], dtype=float)
        pred = np.array([pred_row[c] for c in cols], dtype=float)
        all_human.extend(human)
        all_pred.extend(pred)

        row = {"language": lang, **metrics(human, pred)}
        results.append(row)
        print(f"\n{lang} (n={len(human)})")
        print(f"  Pearson r={row['pearson_r']:.3f} (p={row['pearson_p']:.4f})")
        print(f"  Spearman rho={row['spearman_rho']:.3f} (p={row['spearman_p']:.4f})")
        print(f"  Kendall tau={row['kendall_tau']:.3f} (p={row['kendall_p']:.4f})")
        print(f"  MAE={row['mae']:.3f}  RMSE={row['rmse']:.3f}")

    pooled = {"language": "Pooled (all)", **metrics(np.array(all_human), np.array(all_pred))}
    results.append(pooled)
    print(f"\nPooled (n={pooled['n']})")
    print(f"  Pearson r={pooled['pearson_r']:.3f} (p={pooled['pearson_p']:.4f})")
    print(f"  Spearman rho={pooled['spearman_rho']:.3f} (p={pooled['spearman_p']:.4f})")
    print(f"  Kendall tau={pooled['kendall_tau']:.3f} (p={pooled['kendall_p']:.4f})")
    print(f"  MAE={pooled['mae']:.3f}  RMSE={pooled['rmse']:.3f}")

    fieldnames = ["language", "n", "pearson_r", "pearson_p", "spearman_rho",
                  "spearman_p", "kendall_tau", "kendall_p", "mae", "rmse"]
    print(f"\nWriting results to: {args.output_csv}")
    try:
        args.output_csv.parent.mkdir(parents=True, exist_ok=True)
        with args.output_csv.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(results)
    except OSError as exc:
        raise SystemExit(
            f"Could not write CSV to {args.output_csv}: {exc}\n"
            "(Common causes: the file is open in Excel, or you don't have write "
            "permission to that folder. Try --output-csv <path> to save elsewhere.)"
        )
    print(f"Saved results table: {args.output_csv}")


if __name__ == "__main__":
    main()
